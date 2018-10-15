import csv, sys, time, os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Side, Border
from datetime import datetime

NEW_TICKETS_FILE_NAME = "Results.tsv"
SPREADSHEET_NAME = "Callouts.xlsx"
SHEET_NAME = "Callouts 2018"

try:
    with open(NEW_TICKETS_FILE_NAME) as tsv:
        try:
            wb = load_workbook(SPREADSHEET_NAME)
        except FileNotFoundError:
            print(SPREADSHEET_NAME, 'doesn\'t exist (weekly spreadsheet), please add this file to the directory '
                                    'of this script')
            print('Directory of script:', os.path.dirname(os.path.realpath(__file__)))
            time.sleep(4)
            sys.exit()

        try:
            currentSheet = wb[SHEET_NAME]
        except KeyError:
            print('Sheet named \'' + SHEET_NAME + '\' doesn\'t exist')
            print('Either create a sheet with this name in', SPREADSHEET_NAME, 'or edit script code so it finds a sheet '
                                                                               'that does exist')
            time.sleep(4)
            sys.exit()

        # Find which row to start appending spreadsheet
        alarmColumn = currentSheet['A']
        print('Calculating where to append spreadsheet')

        startAppending = False
        i = 1    # Start at 1 as first result will always be 'column header'
        while not startAppending:
            if alarmColumn[i].value is None:
                startAppending = True
                startingRowNumber = i + 1
            i += 1

        print('Will append spreadsheet from row #' + str(startingRowNumber))

        # Skips first row of TSV file due to column headers
        iterResults = iter(csv.reader(tsv, dialect="excel-tab"))
        next(iterResults)

        # Extracting data from TSV file
        i = 0
        nagiosFiller = "Nagios issued"
        hostFiller = "on_host_"
        for row in iterResults:
            hostStart = None
            service = None
            hostCheck = True
            workingHours = False
            ticketID = int(row[0])
            
            ticketCreated = datetime.strptime(row[15], '%Y-%m-%d %H:%M:%S')
            #Excel date time counts days and seconds since 1900-01-00 but mistakenly treats 1900 as a leap year
            excelTicketCreated = ticketCreated - datetime(1899, 12, 30)
            dateCreated = float(excelTicketCreated.days)
            timeCreated = float(excelTicketCreated.seconds) / 86400

            # Is callout in work hours (8:30-5:00 converted into percent of day)?
            weekday = ticketCreated.isoweekday()
            if (60*8+30)/(24*60) < timeCreated < 17/24 and weekday < 6:
                workingHours = True

            # Get Nagios alarm (or subject if not from Nagios)
            alarm = row[2]
            if nagiosFiller in alarm:
                for k in range(len(nagiosFiller), len(alarm)):
                    if 'T' in alarm[k]:
                        alarm = alarm[k:]
                        break

            # Service
            if 'ceph' in alarm.lower():
                service = 'CEPH'
            elif 'arc-ce' in alarm.lower():
                service = 'CE'
            elif 'gdss' in alarm.lower():
                service = 'DISK Server'
            elif 'fts' in alarm.lower():
                service = 'FTS'

            # Get hostname from Nagios alarm
            if hostFiller in alarm:
                j = len(alarm) - 1
                while hostCheck:
                    if alarm[j] == '_':
                        hostStart = j
                        hostCheck = False
                    j -= 1
                hostname = alarm[hostStart + 1:]

            # Putting data into spreadsheet
            currentRow = startingRowNumber + i
            currentSheet.cell(row=currentRow, column=1, value=alarm)
            if hostStart is not None:
                currentSheet.cell(row=currentRow, column=2, value=hostname)
            currentSheet.cell(row=currentRow, column=3, value=dateCreated)
            currentSheet.cell(row=currentRow, column=4, value=timeCreated)
            if service is not None:
                currentSheet.cell(row=currentRow, column=6, value=service)
            # RT query
            currentSheet.cell(row=currentRow, column=5, value=ticketID)
            currentSheet.cell(row=currentRow, column=5).hyperlink = 'https://helpdesk.gridpp.rl.ac.uk/Ticket/Display' \
                                                                    '.html?id=' + str(ticketID)
            if workingHours:
                currentSheet.cell(row=currentRow, column=8, value='N/A')
                currentSheet.cell(row=currentRow, column=10, value='Work hours')
            i += 1

        # Merge cells
        currentDate = datetime.now().strftime('%d-%b')
        currentSheet.cell(row=startingRowNumber, column=12, value=currentDate)
        currentSheet.merge_cells(start_row=startingRowNumber, start_column=12, end_row=currentRow, end_column=12)

        # Setting inner borders and cell alignment
        rows = currentSheet.iter_rows(min_row=startingRowNumber, min_col=1, max_row=currentRow, max_col=17)
        innerBorderStyle = Side(border_style='thin', color='FF000000')
        innerBorderFormat = Border(left=innerBorderStyle, right=innerBorderStyle, top=innerBorderStyle,
                                   bottom=innerBorderStyle)
        for row in rows:
            for cell in row:
                if cell.column == 'A' or cell.column == 'K':
                    # These columns need to be horizontally left aligned (alarm and comment columns)
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                cell.border = innerBorderFormat

        # Setting outer border
        # Code found at: https://stackoverflow.com/questions/34520764/apply-border-to-range-of-cells-using-openpyxl
        # Written by Yaroslav Admin, edited by Adam Stewart
        outerRows = currentSheet.iter_rows(min_row=startingRowNumber, min_col=1, max_row=currentRow, max_col=17)
        outerBorderStyle = Side(border_style='medium', color='FF000000')
        outerRows = list(outerRows)
        max_y = len(outerRows) - 1
        for pos_y, cells in enumerate(outerRows):
            max_x = len(cells) - 1  # index of the last cell
            for pos_x, cell in enumerate(cells):
                border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom)

                # Checking if an edge cell
                if pos_x == 0:
                    border.left = outerBorderStyle
                if pos_x == max_x:
                    border.right = outerBorderStyle
                if pos_y == 0:
                    border.top = outerBorderStyle
                if pos_y == max_y:
                    border.bottom = outerBorderStyle

                # Set new border only if it's one of the edge cells
                if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                    cell.border = border

        wb.save(SPREADSHEET_NAME)
        print('Spreadsheet changes saved')
except FileNotFoundError:
    print(NEW_TICKETS_FILE_NAME, '(this week\'s tickets) not found, please add to the same directory of the script')
    print('Directory of script:', os.path.dirname(os.path.realpath(__file__)))
    time.sleep(4)
    sys.exit()
